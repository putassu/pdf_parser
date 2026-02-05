import openpyxl
from openpyxl.utils import get_column_letter
import logging
import time
import base64
import io

# ================= ГЛОБАЛЬНЫЕ НАСТРОЙКИ ПО УМОЛЧАНИЮ =================
DEFAULT_PARAMS = {
    "H_TOLERANCE": 1,
    "V_TOLERANCE": 1,
    "MAX_CHARS_BLOCK": 3000,
    "MIN_TABLE_ROWS": 10,
    "VALIDATION_THRESHOLD": 0.98,
    "SHOW_MERGED_MAP": True
}

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger("SeniorAnalyzer")

def mock_do_vlm(img_data_base64):
    """Моковая функция VLM: возвращает первые 50 символов base64."""
    return f"VLM_ANALYSIS: {img_data_base64[:50]}..."

class RobustExcelParser:
    def __init__(self, **kwargs):
        self.params = DEFAULT_PARAMS.copy()
        self.params.update(kwargs)

    def update_params(self, **kwargs):
        self.params.update(kwargs)

    def _image_to_base64(self, img):
        """Конвертация объекта изображения openpyxl в base64 string."""
        try:
            # Картинка лежит в img.ref (объект изображения)
            # В разных версиях openpyxl доступ может отличаться
            from PIL import Image
            image = Image.open(img.ref)
            buffered = io.BytesIO()
            image.save(buffered, format="PNG")
            return base64.b64encode(buffered.getvalue()).decode('utf-8')
        except Exception as e:
            # Если не получилось через PIL, пробуем прочитать байты напрямую, если это возможно
            try:
                raw_data = img._data() # Внутренний метод некоторых версий
                return base64.b64encode(raw_data).decode('utf-8')
            except:
                return f"ERROR_LOADING_IMAGE: {str(e)}"

    def parse_file(self, file_path: str, vlm_callback=None):
        start_time = time.time()
        logger.info(f"Processing: {file_path}")
        
        try:
            # Открываем для чтения данных и стилей
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return {"error": str(e)}

        manifest = []
        for sheet in wb.worksheets:
            sheet_report = self._process_sheet(sheet, vlm_callback)
            if sheet_report:
                manifest.append({
                    "sheet_name": sheet.title,
                    "data": sheet_report
                })

        logger.info(f"Parsing finished in {time.time() - start_time:.2f}s")
        return manifest

    def _is_significant(self, cell, in_merged):
        if cell.value is not None and str(cell.value).strip() != "":
            return True
        if in_merged: return True
        b = cell.border
        if b and (b.left.style or b.right.style or b.top.style or b.bottom.style):
            return True
        if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
            return True
        return False

    def _process_sheet(self, sheet, vlm_callback):
        sig_data = {}
        merged_lookup = {}
        
        for rng in sheet.merged_cells.ranges:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    merged_lookup[(r, c)] = rng

        for row in sheet.iter_rows():
            for cell in row:
                if self._is_significant(cell, (cell.row, cell.column) in merged_lookup):
                    sig_data[(cell.row, cell.column)] = cell.value

        if not sig_data: return None

        # Обработка изображений
        images_results = []
        if hasattr(sheet, '_images') and sheet._images:
            for img in sheet._images:
                b64_str = self._image_to_base64(img)
                analysis = vlm_callback(b64_str) if vlm_callback else "No VLM callback"
                
                # Вместо объекта Anchor сохраняем строку, чтобы не было ошибки repr
                anchor_pos = "Unknown"
                if hasattr(img, 'anchor'):
                    # Обычно это объект TwoCellAnchor или OneCellAnchor
                    try:
                        anchor_pos = f"{get_column_letter(img.anchor._from.col + 1)}{img.anchor._from.row + 1}"
                    except:
                        anchor_pos = str(img.anchor)

                images_results.append({
                    "anchor": anchor_pos,
                    "vlm_res": analysis
                })

        # Кластеризация
        clusters = self._cluster_regions(list(sig_data.keys()))
        
        regions = []
        cells_covered = 0
        for cluster in clusters:
            cells_covered += len(cluster)
            reg = self._analyze_region(sheet, cluster, sig_data, merged_lookup)
            regions.append(reg)

        coverage = cells_covered / len(sig_data)
        if coverage < self.params["VALIDATION_THRESHOLD"]:
            logger.warning(f"Low coverage ({coverage:.2%}) on {sheet.title}")

        return {
            "regions": regions, 
            "images": images_results, 
            "coverage": coverage
        }

    def _cluster_regions(self, coords):
        coords_set = set(coords)
        visited = set()
        clusters = []
        sorted_coords = sorted(coords)

        for node in sorted_coords:
            if node in visited: continue
            cluster, queue = [], [node]
            visited.add(node)
            while queue:
                r, c = queue.pop(0)
                cluster.append((r, c))
                for nr in range(r - self.params["V_TOLERANCE"], r + self.params["V_TOLERANCE"] + 1):
                    for nc in range(c - self.params["H_TOLERANCE"], c + self.params["H_TOLERANCE"] + 1):
                        neighbor = (nr, nc)
                        if neighbor in coords_set and neighbor not in visited:
                            visited.add(neighbor); queue.append(neighbor)
            clusters.append(cluster)
        return clusters

    def _analyze_region(self, sheet, cluster_coords, sig_data, merged_lookup):
        rows, cols = [c[0] for c in cluster_coords], [c[1] for c in cluster_coords]
        min_r, max_r, min_c, max_c = min(rows), max(rows), min(cols), max(cols)
        
        lines = []
        for r in range(min_r, max_r + 1):
            row_parts, skip_cols = [], set()
            for c in range(min_c, max_c + 1):
                if c in skip_cols: continue
                coord = (r, c)
                if coord in merged_lookup:
                    m = merged_lookup[coord]
                    if r == m.min_row and c == m.min_col:
                        val = sig_data.get(coord, "")
                        addr = str(m) if self.params["SHOW_MERGED_MAP"] else f"{get_column_letter(c)}{r}"
                        row_parts.append(f"[{addr}]: {str(val).strip()}")
                    for col_idx in range(m.min_col, m.max_col + 1): skip_cols.add(col_idx)
                else:
                    val = sig_data.get(coord)
                    if val is not None:
                        row_parts.append(f"[{get_column_letter(c)}{r}]: {str(val).strip()}")
            if row_parts: lines.append(" | ".join(row_parts))

        is_table = len("\n".join(lines)) > self.params["MAX_CHARS_BLOCK"] and len(lines) > self.params["MIN_TABLE_ROWS"]
        if is_table:
            preview = "\n".join(lines[:5]) + f"\n... [SKIPPED {len(lines)-10} ROWS] ...\n" + "\n".join(lines[-5:])
            res_type = "data_table"
        else:
            preview = "\n".join(lines); res_type = "data_form"

        return {
            "type": res_type,
            "range": f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}",
            "preview": preview
        }

# --- КРАСИВЫЙ ВЫВОД И ТЕСТ ---

if __name__ == "__main__":
    # Можно менять параметры прямо здесь
    parser = RobustExcelParser(V_TOLERANCE=1, H_TOLERANCE=1)
    
    file_path = r"C:\\Users\\sigur\\docling\\xlsx_parser\\Анкета_соискателя_на_вакантную_должность_1_1 (3) — копия (2).xlsx"
    results = parser.parse_file(file_path, vlm_callback=mock_do_vlm)

    for sheet_data in results:
        s_name = sheet_data['sheet_name']
        data = sheet_data['data']
        
        print(f"\n{'='*30} SHEET: {s_name} (Cov: {data['coverage']:.2%}) {'='*30}")
        
        # Вывод регионов
        for r in data['regions']:
            print(f"\n--- Region ({r['type']}) {r['range']} ---")
            print(r['preview'])
            print("-" * 60)
            
        # Вывод изображений
        if data['images']:
            print(f"\nFound {len(data['images'])} image(s):")
            for img in data['images']:
                print(f"  • Position {img['anchor']}: {img['vlm_res']}")

