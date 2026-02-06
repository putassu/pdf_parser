import openpyxl
from openpyxl.utils import get_column_letter
import logging
import os
import uuid
import asyncio
import shutil
from typing import Dict, List, Optional, Callable, Any

# Импортируем настройки по умолчанию
try:
    from config import DEFAULT_SETTINGS
except ImportError:
    # Запасной вариант, если config.py не найден
    DEFAULT_SETTINGS = {
        "H_TOLERANCE": 1,
        "V_TOLERANCE": 1,
        "MAX_CHARS_BLOCK": 3000,
        "MIN_TABLE_ROWS": 10,
        "VALIDATION_THRESHOLD": 0.98,
        "SHOW_MERGED_MAP": True
    }

logger = logging.getLogger("ExcelAnalyzer")

class RobustExcelParser:
    def __init__(self, global_config: Optional[Dict] = None):
        """
        Инициализация парсера.
        :param global_config: Глобальные настройки. Если None, берутся из config.py
        """
        self.global_config = global_config or DEFAULT_SETTINGS.copy()
        self.sheet_configs: Dict[str, Dict] = {}
        self.temp_dir = "temp_vlm_images"
        
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)

    def set_sheet_config(self, sheet_name: str, params: Dict):
        """Установка индивидуальных параметров для конкретного листа."""
        if sheet_name not in self.sheet_configs:
            self.sheet_configs[sheet_name] = self.global_config.copy()
        self.sheet_configs[sheet_name].update(params)
        logger.info(f"Custom config set for sheet: {sheet_name}")

    def _get_params(self, sheet_name: str) -> Dict:
        """Получение параметров для листа (индивидуальных или глобальных)."""
        return self.sheet_configs.get(sheet_name, self.global_config)

    def _is_significant(self, cell, in_merged: bool) -> bool:
        """Проверка ячейки на наличие данных, границ или заливки."""
        if cell.value is not None and str(cell.value).strip() != "":
            return True
        if in_merged:
            return True
        # Проверка границ
        b = cell.border
        if b and any([b.left.style, b.right.style, b.top.style, b.bottom.style]):
            return True
        # Проверка заливки
        if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
            return True
        return False

    async def parse_file(
        self, 
        file_path: str, 
        target_sheets: Optional[List[str]] = None, 
        image_callback: Optional[Callable] = None,
        session: Any = None
    ) -> Dict:
        """
        Основной метод парсинга файла.
        :param vlm_callback: Асинхронная функция вида func(session, file_path, filename)
        :param target_sheets: Список листов для обработки. Если None - все.
        """
        logger.info(f"Opening workbook: {file_path}")
        try:
            # data_only=True позволяет получать значения формул
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return {"error": str(e)}

        manifest = {}
        
        # Определяем список листов для обработки
        worksheets = wb.worksheets
        if target_sheets:
            worksheets = [s for s in wb.worksheets if s.title in target_sheets]

        for sheet in worksheets:
            params = self._get_params(sheet.title)
            logger.info(f"Processing sheet: {sheet.title} with params: {params}")
            
            sheet_data = await self._process_sheet(sheet, params, image_callback, session)
            if sheet_data:
                manifest[sheet.title] = sheet_data

        # Очистка временной папки после всего процесса
        self._cleanup_temp()
        
        return manifest

    async def _process_sheet(self, sheet, params, image_callback, session) -> Dict:
        sig_data = {}
        merged_lookup = {}
        
        # 1. Кэшируем объединенные ячейки
        for rng in sheet.merged_cells.ranges:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    merged_lookup[(r, c)] = rng

        # 2. Собираем все значимые координаты
        for row in sheet.iter_rows():
            for cell in row:
                if self._is_significant(cell, (cell.row, cell.column) in merged_lookup):
                    sig_data[(cell.row, cell.column)] = cell.value

        if not sig_data:
            return {"status": "empty"}

        # 3. Обработка изображений (если есть)
        vlm_results = []
        if hasattr(sheet, '_images') and sheet._images:
            for idx, img in enumerate(sheet._images):
                vlm_res = await self._handle_image(img, sheet.title, image_callback, session)
                if vlm_res:
                    vlm_results.append(vlm_res)

        # 4. Кластеризация
        clusters = self._cluster_regions(list(sig_data.keys()), params)
        
        # 5. Анализ регионов
        regions = []
        cells_covered = 0
        for cluster in clusters:
            cells_covered += len(cluster)
            region_report = self._analyze_region(sheet, cluster, sig_data, merged_lookup, params)
            regions.append(region_report)

        coverage = cells_covered / len(sig_data) if sig_data else 1.0
        
        return {
            "params_used": params,
            "coverage": round(coverage, 4),
            "regions": regions,
            "vlm_data": vlm_results
        }

    async def _handle_image(self, img, sheet_title, image_callback, session) -> Optional[Dict]:
        """Сохранение изображения и вызов VLM."""
        unique_id = str(uuid.uuid4())[:8]
        filename = f"{sheet_title}_{unique_id}.png".replace(" ", "_")
        file_path = os.path.join(self.temp_dir, filename)
        
        try:
            # Извлечение изображения
            from PIL import Image as PILImage
            # У openpyxl изображение хранится в img.ref (это BytesIO или путь)
            image_data = PILImage.open(img.ref)
            image_data.save(file_path)

            analysis = "VLM Not Configured"
            if image_callback:
                # Вызываем переданную асинхронную функцию
                analysis = await image_callback(session, file_path, filename)
            
            # Определяем позицию (якорь)
            anchor = "Unknown"
            try:
                anchor = f"{get_column_letter(img.anchor._from.col + 1)}{img.anchor._from.row + 1}"
            except: pass

            return {"anchor": anchor, "vlm_analysis": analysis}
        except Exception as e:
            logger.error(f"Error processing image on {sheet_title}: {e}")
            return None
        finally:
            # Удаляем конкретный файл сразу после обработки
            if os.path.exists(file_path):
                os.remove(file_path)

    def _cluster_regions(self, coords: List[tuple], params: Dict) -> List[List[tuple]]:
        """Алгоритм поиска связанных компонентов на основе допусков."""
        coords_set = set(coords)
        visited = set()
        clusters = []
        sorted_coords = sorted(coords)

        v_tol = params["V_TOLERANCE"]
        h_tol = params["H_TOLERANCE"]

        for node in sorted_coords:
            if node in visited:
                continue
            
            cluster = []
            queue = [node]
            visited.add(node)
            
            while queue:
                r, c = queue.pop(0)
                cluster.append((r, c))
                
                # Поиск соседей в окне допусков
                for nr in range(r - v_tol, r + v_tol + 1):
                    for nc in range(c - h_tol, c + h_tol + 1):
                        neighbor = (nr, nc)
                        if neighbor in coords_set and neighbor not in visited:
                            visited.add(neighbor)
                            queue.append(neighbor)
            clusters.append(cluster)
        return clusters

    def _analyze_region(self, sheet, cluster_coords, sig_data, merged_lookup, params) -> Dict:
        """Анализирует блок ячеек и превращает его в текст."""
        rows = [c[0] for c in cluster_coords]
        cols = [c[1] for c in cluster_coords]
        min_r, max_r, min_c, max_c = min(rows), max(rows), min(cols), max(cols)
        
        lines = []
        for r in range(min_r, max_r + 1):
            row_parts = []
            skip_cols = set()
            for c in range(min_c, max_c + 1):
                if c in skip_cols: continue
                
                coord = (r, c)
                if coord in merged_lookup:
                    m = merged_lookup[coord]
                    if r == m.min_row and c == m.min_col:
                        val = sig_data.get(coord, "")
                        addr = str(m) if params["SHOW_MERGED_MAP"] else f"{get_column_letter(c)}{r}"
                        row_parts.append(f"[{addr}]: {str(val).strip()}" if val is not None else f"[{addr}]: ")
                    # Пропускаем остальные колонки в этом объединении для текущей строки
                    for col_idx in range(m.min_col, m.max_col + 1):
                        skip_cols.add(col_idx)
                else:
                    val = sig_data.get(coord)
                    if val is not None:
                        row_parts.append(f"[{get_column_letter(c)}{r}]: {str(val).strip()}")
            
            if row_parts:
                lines.append(" | ".join(row_parts))

        # Определение типа региона
        full_text = "\n".join(lines)
        is_table = len(full_text) > params["MAX_CHARS_BLOCK"] and len(lines) > params["MIN_TABLE_ROWS"]
        
        if is_table:
            head = lines[:5]
            tail = lines[-5:]
            skipped = len(lines) - 10
            preview = "\n".join(head) + f"\n... [SKIPPED {skipped} ROWS] ...\n" + "\n".join(tail)
            r_type = "data_table"
        else:
            preview = full_text
            r_type = "data_form"

        return {
            "type": r_type,
            "range": f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}",
            "preview": preview,
            "metrics": {"rows": len(lines), "cells": len(cluster_coords)}
        }

    def _cleanup_temp(self):
        """Полная очистка временной папки."""
        try:
            if os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
                os.makedirs(self.temp_dir)
        except Exception as e:
            logger.error(f"Cleanup error: {e}")

# --- Пример использования (Mock-тест) ---
async def main():
    # Имитация VLM функции
    async def mock_vlm(session, file_path, filename):
        await asyncio.sleep(0.1) # имитация сетевой задержки
        return f"VLM_RESULT_FOR_{filename}"

    parser = RobustExcelParser()
    
    # 1. Первый проход
    results = await parser.parse_file("Анкета.xlsx", image_callback=mock_vlm)
    
    # 2. Имитация команды от LLM: "Перепарсить Лист1 с новыми настройками"
    parser.set_sheet_config("Лист1", {"V_TOLERANCE": 0, "H_TOLERANCE": 0})
    
    # 3. Перепаршиваем только конкретный лист
    updated_results = await parser.parse_file("Анкета.xlsx", target_sheets=["Лист1"], image_callback=mock_vlm)
    
    # Обновляем основные результаты
    results.update(updated_results)
    
    print("Парсинг завершен. Количество листов:", len(results))

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    asyncio.run(main())
